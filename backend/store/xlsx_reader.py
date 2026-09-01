"""
Reading spreadsheets that arrive from the outside world — Phase C1.4.

WHY THIS IS NOT `openpyxl.load_workbook(f)`
-------------------------------------------
The owner's own product workbook does not open. openpyxl raises

    ValueError: Value must be one of {'warning', 'stop', 'information'}

because Google Sheets exported twenty `<dataValidation>` nodes carrying
`errorStyle="error"`, and `error` is not in the OOXML enumeration. The
validations are dropdown hints for whoever types into the sheet; they have
nothing to do with the data. Refusing the file over them would mean telling the
owner their own template is invalid — technically defensible, useless in a shop.

So the reader takes a copy IN MEMORY, removes only the `<dataValidation>` and
`<dataValidations>` elements from the worksheet parts, and hands the result to
openpyxl. Namespace-aware parsing, not a regex over XML: `dataValidation` also
appears inside `<extLst>` blocks and in worksheet text, and a regex that ate the
wrong bytes would corrupt a file we are about to base stock decisions on.

The original file is never modified — it is never even opened for writing.

WHAT ELSE IS DEFENDED AGAINST
-----------------------------
An uploaded spreadsheet is untrusted input that arrives as a ZIP full of XML.
Before anything is parsed:

  · the extension must be `.xlsx` — `.xlsm` carries macros and is refused
  · the upload is size-capped, compressed AND decompressed
  · the entry count and compression ratio are capped (a "zip bomb" is a 1 MB
    upload that becomes 40 GB in RAM)
  · `vbaProject.bin` anywhere in the archive is a refusal
  · formulas are NEVER evaluated and cached results are NEVER trusted

That last one is a data-integrity rule, not a security one, and it is in §17:
a mapped cell containing a formula is a row ERROR telling the operator to
convert it to a value. The alternative is importing whatever number Excel
happened to leave in the file, which may have been computed from rows that are
not in the upload.
"""

from __future__ import annotations

import io
import zipfile
from xml.etree import ElementTree as ET

# --- limits (§16) ---------------------------------------------------------
MAX_UPLOAD_BYTES = 10 * 1024 * 1024          # 10 MB compressed
MAX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024   # 200 MB expanded
MAX_ZIP_ENTRIES = 512
MAX_COMPRESSION_RATIO = 200
MAX_DATA_ROWS = 5000                          # per job
MAX_COLUMNS = 128
SAMPLE_ROWS = 25                              # what an inspection needs to see

# Two ways to read a sheet, and confusing them is how a file gets half-imported.
#
#   SAMPLE       "show me what this file looks like".  Stopping early is the
#                POINT, and the caller must present it as a sample.
#   FULL_IMPORT  "read this file so it can be applied".  Stopping early is a
#                CORRECTNESS FAILURE: the rows past the cut would silently not
#                be imported while the job still reported zero errors.
#
# So FULL_IMPORT never truncates. It reads one row PAST the limit purely to
# find out whether the limit was exceeded, and then refuses the whole file.
SAMPLE = 'sample'
FULL_IMPORT = 'full_import'

SPREADSHEET_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


class XlsxError(Exception):
    """The upload cannot be read as a spreadsheet this system will accept."""


class XlsxTooLarge(XlsxError):
    pass


class XlsxTooManyRows(XlsxError):
    """More data rows than one job may import. Refused whole, never trimmed."""


class FormulaCell(Exception):
    """A mapped cell holds a formula. Never evaluated, never guessed."""


# --------------------------------------------------------------------------
# Validation of the container
# --------------------------------------------------------------------------

def check_upload(uploaded, *, filename: str | None = None) -> bytes:
    """
    Validate the upload envelope and return its bytes.

    Everything here happens before a single XML byte is parsed, because the
    cheapest way to survive a malicious archive is not to open it.
    """
    name = (filename or getattr(uploaded, 'name', '') or '').strip()
    lowered = name.lower()
    if not lowered.endswith('.xlsx'):
        if lowered.endswith(('.xlsm', '.xls')):
            raise XlsxError(
                'Sólo se aceptan archivos .xlsx. Un .xlsm puede contener macros '
                'y un .xls es un formato antiguo distinto. Guarda el archivo '
                'como .xlsx y vuelve a subirlo.'
            )
        raise XlsxError('Sólo se aceptan archivos .xlsx.')

    data = uploaded.read() if hasattr(uploaded, 'read') else bytes(uploaded)
    if len(data) > MAX_UPLOAD_BYTES:
        raise XlsxTooLarge(
            f'El archivo pesa {len(data) // (1024 * 1024)} MB y el límite es '
            f'{MAX_UPLOAD_BYTES // (1024 * 1024)} MB.'
        )
    if not data:
        raise XlsxError('El archivo está vacío.')

    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        raise XlsxError(
            'El archivo no es un .xlsx válido. Si lo exportaste de otro '
            'programa, vuelve a guardarlo como "Libro de Excel (.xlsx)".'
        ) from None

    infos = archive.infolist()
    if len(infos) > MAX_ZIP_ENTRIES:
        raise XlsxTooLarge('El archivo contiene demasiadas partes internas.')

    total = 0
    for info in infos:
        lowered_entry = info.filename.lower()
        if lowered_entry.endswith('vbaproject.bin') or '/vba' in lowered_entry:
            raise XlsxError(
                'El archivo contiene macros. Guarda una copia como .xlsx sin '
                'macros y vuelve a subirla.'
            )
        total += info.file_size
        if total > MAX_UNCOMPRESSED_BYTES:
            raise XlsxTooLarge(
                'El archivo se expande a un tamaño desproporcionado y no se '
                'procesará.'
            )
        if info.compress_size and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
            raise XlsxTooLarge(
                'El archivo se expande a un tamaño desproporcionado y no se '
                'procesará.'
            )

    if not any(i.filename == 'xl/workbook.xml' for i in infos):
        raise XlsxError('El archivo no contiene una hoja de cálculo legible.')

    return data


# --------------------------------------------------------------------------
# Sanitising the parts openpyxl chokes on
# --------------------------------------------------------------------------

def _strip_data_validations(sheet_xml: bytes) -> tuple[bytes, bool]:
    """
    Remove `<dataValidation(s)>` from one worksheet part.

    Returns `(xml, changed)`. Parsed with ElementTree so the removal is
    structural: an element is dropped because it IS a dataValidation node in the
    spreadsheetml namespace, not because its name appeared in the bytes.
    """
    try:
        ET.register_namespace('', SPREADSHEET_NS)
        root = ET.fromstring(sheet_xml)
    except ET.ParseError:
        # Unparseable XML is not something to paper over — hand the original
        # back and let openpyxl produce its own diagnosis.
        return sheet_xml, False

    changed = False
    targets = (
        f'{{{SPREADSHEET_NS}}}dataValidation',
        f'{{{SPREADSHEET_NS}}}dataValidations',
    )
    # Walk parents, because ElementTree has no parent pointers and a node can
    # sit at the sheet root or nested inside an extension list.
    for parent in root.iter():
        doomed = [child for child in list(parent) if child.tag in targets]
        for child in doomed:
            parent.remove(child)
            changed = True

    if not changed:
        return sheet_xml, False
    return ET.tostring(root, encoding='UTF-8', xml_declaration=True), True


def sanitized_workbook_bytes(data: bytes) -> tuple[bytes, list[str]]:
    """
    Return a repaired copy of the workbook plus a list of what was repaired.

    The repairs are reported, not silent: an operator who is told "twenty
    dropdown validations were ignored" can tell that from "the file was
    quietly changed", and only the first of those is true.
    """
    source = zipfile.ZipFile(io.BytesIO(data))
    notes: list[str] = []
    out = io.BytesIO()
    touched = 0

    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if (info.filename.startswith('xl/worksheets/')
                    and info.filename.endswith('.xml')):
                payload, changed = _strip_data_validations(payload)
                if changed:
                    touched += 1
            target.writestr(info.filename, payload)

    if touched:
        notes.append(
            f'Se ignoraron las validaciones de datos (listas desplegables) de '
            f'{touched} hoja(s). Son ayudas para escribir en la plantilla y no '
            f'afectan a los valores importados.'
        )
    return out.getvalue(), notes


def has_data_validations(data: bytes) -> bool:
    """Whether any worksheet part declares a `<dataValidation>`."""
    archive = zipfile.ZipFile(io.BytesIO(data))
    needle = b'dataValidation'
    for info in archive.infolist():
        if (info.filename.startswith('xl/worksheets/')
                and info.filename.endswith('.xml')
                and needle in archive.read(info.filename)):
            return True
    return False


def load_workbook(data: bytes):
    """
    Open the workbook, stripping data validations first if it has any.

    WHY THE STRIP IS UNCONDITIONAL RATHER THAN A FALLBACK
    -----------------------------------------------------
    The obvious shape is "try to open it; repair only if that fails". It does
    not work here, and the reason is worth writing down: openpyxl in read-only
    mode parses worksheets LAZILY. `load_workbook()` on the owner's file
    succeeds. The `ValueError` about `errorStyle` surfaces later, from
    `iter_rows()`, halfway through the caller's loop — so a fallback wrapped
    around the open never runs, and the failure arrives somewhere that cannot
    recover from it.

    Data validations are dropdown hints for whoever types into the sheet. No
    importer needs them. Removing them when present is deterministic, costs one
    in-memory rewrite, and does not depend on which openpyxl version rejects
    which attribute.
    """
    import openpyxl

    notes: list[str] = []
    payload = data
    if has_data_validations(data):
        payload, notes = sanitized_workbook_bytes(data)

    try:
        return openpyxl.load_workbook(
            io.BytesIO(payload), read_only=True, data_only=False,
        ), notes
    except Exception as exc:
        raise XlsxError(
            f'No se pudo leer el archivo de Excel: {exc}. Ábrelo y vuelve a '
            f'guardarlo como .xlsx.'
        ) from exc


# --------------------------------------------------------------------------
# Reading values
# --------------------------------------------------------------------------

def sheet_names(workbook) -> list[str]:
    return list(workbook.sheetnames)


def _cell_value(cell):
    """
    The value of one cell, refusing formulas.

    `data_only=False` means openpyxl hands back the formula STRING for a
    formula cell rather than Excel's cached result. That is the point: the
    cached result was computed by whatever machine last saved the file, from
    rows that may not be in this upload. §17 — a mapped formula is a row error.
    """
    value = cell.value
    if isinstance(value, str) and value.startswith('='):
        raise FormulaCell(value)
    return value


def normalize_scalar(value):
    """
    Turn one cell into a stable string, without inventing anything.

    THE INTEGER-THAT-IS-REALLY-A-CODE PROBLEM (§29)
    ----------------------------------------------
    Excel stores `CODIGO EAN` as a NUMBER. Python then hands us `310000000001.0`
    or, worse, `3.10000000001e+11`, and `str()` on either produces something no
    barcode reader would recognise. A float that is mathematically an exact
    integer is rendered as that integer, in full decimal notation.

    What this does NOT do is add leading zeros. If Excel already dropped a
    leading zero, that information is gone from the file, and reconstructing it
    would mean inventing a digit that determines which physical product this is.
    The caller warns instead.
    """
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'SI' if value else 'NO'
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    from datetime import date, datetime
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def read_rows(workbook, sheet_name: str, *, header_row: int = 1,
              limit: int = MAX_DATA_ROWS, mode: str = FULL_IMPORT):
    """
    Read one sheet as `(headers, rows, truncated)`.

    `headers` is the normalised header row. `rows` is a list of
    `(row_number, values, numeric_columns)`:

      · `row_number`     the line the OPERATOR sees in Excel. Off-by-one here
                         means they go and fix the wrong row.
      · `values`         `{column_index: text}` for the cells that had content.
      · `numeric_columns` the subset of those whose SOURCE CELL was a number.
                         Kept because "this looks like digits" and "Excel stored
                         this as a number" are different facts, and only the
                         second one justifies warning that a leading zero may
                         have been eaten.

    `truncated` is meaningful ONLY in SAMPLE mode, where it means "there is more
    of this file". In FULL_IMPORT mode it is always False, because exceeding the
    limit raises `XlsxTooManyRows` instead — see the note on the mode constants.

    THE OFF-BY-ONE THAT MATTERS FOR THE LIMIT
    -----------------------------------------
    Blank rows are collected while reading (a blank row between records is
    normal and must keep its number) and only trimmed from the END afterwards.
    So the limit is applied AFTER trimming: a sheet with 5000 records followed
    by 200 rows of leftover template formatting is a 5000-row file, not a
    5200-row one, and must not be refused for rows that contain nothing.
    """
    if sheet_name not in workbook.sheetnames:
        raise XlsxError(f'La hoja «{sheet_name}» no existe en el archivo.')
    if mode not in (SAMPLE, FULL_IMPORT):
        raise XlsxError(f'Modo de lectura desconocido: {mode}')
    sheet = workbook[sheet_name]

    # One row past the limit, so "exactly at the limit" and "over it" are
    # distinguishable. In SAMPLE mode the extra row is what proves there is more.
    hard_stop = limit + 1

    headers: list[str] = []
    rows: list[tuple[int, dict[int, object], set[int]]] = []

    previous = 0
    for row in sheet.iter_rows(min_row=1, max_col=MAX_COLUMNS):
        if not row:
            continue
        # In read-only mode openpyxl yields `EmptyCell` for blank cells, and an
        # EmptyCell has no `.row`. A fully blank row — §24 says skip it, not
        # fail — is therefore all EmptyCells, and asking the first one for its
        # row number raises. Take the number from whichever cell has one, and
        # fall back to counting, so the number an operator is shown still
        # matches the line they see in Excel.
        number = next(
            (c.row for c in row if getattr(c, 'row', None) is not None),
            previous + 1,
        )
        previous = number
        if number < header_row:
            continue
        if number == header_row:
            headers = [normalize_scalar(c.value) for c in row]
            # Trailing empties carry no meaning and would become phantom columns.
            while headers and not headers[-1]:
                headers.pop()
            continue

        if len(rows) >= hard_stop:
            break

        values: dict[int, object] = {}
        numeric: set[int] = set()
        for index, cell in enumerate(row):
            try:
                raw = _cell_value(cell)
            except FormulaCell:
                values[index] = FORMULA
                continue
            if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                numeric.add(index)
            text = normalize_scalar(raw)
            if text != '':
                values[index] = text
            else:
                numeric.discard(index)
        rows.append((number, values, numeric))

    # Trailing all-blank rows are an artefact of the template's formatting (the
    # owner's product sheet declares 202 rows and contains 0), not data.
    while rows and not rows[-1][1]:
        rows.pop()

    if len(rows) > limit:
        if mode == FULL_IMPORT:
            # Refused WHOLE. Trimming to the limit would stage 5000 rows, report
            # zero errors and apply cleanly, while the rest of the file silently
            # never arrived — a half-imported catalogue that looks complete.
            raise XlsxTooManyRows(
                f'La hoja «{sheet_name}» tiene más de {limit} filas de datos. '
                f'Divide el archivo en partes de {limit} filas o menos y súbelas '
                f'por separado: no se importa un archivo a medias.'
            )
        return headers, rows[:limit], True

    return headers, rows, False


class _Formula:
    """Marker for a formula cell: distinguishable from any legitimate value."""

    def __repr__(self):
        return '<FORMULA>'

    def __bool__(self):
        return True


FORMULA = _Formula()
