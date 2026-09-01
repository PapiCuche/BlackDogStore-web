"""
Files the platform hands BACK — Phase C1.4.

Two of them, and the second is the one that matters operationally:

  · a blank product template, so a new tenant is not asked to guess the columns;
  · the current inventory, so counting stock is "download, walk the shelves,
    type, upload" instead of transcribing six hundred lines by hand.

THE COUNT SHEET IS OFFERED BLANK ON PURPOSE
-------------------------------------------
Downloading the inventory with the current quantities already filled in is
convenient and is also how a physical count stops being a count: whoever holds
the sheet reads "14", sees fourteen-ish, and writes nothing. A blank column
forces an actual look at the shelf. Both are offered because pre-filled is right
for correcting a few known lines, and blank is right for a real stocktake — but
the operator picks, deliberately.

WHAT THE EXPORT MUST NOT DO
---------------------------
It must not emit `ALMACEN 1 - 11416`. That header belongs to the system that
produced the owner's file, and the number in it is that system's warehouse id.
Writing it for every tenant would bake one customer's data into everybody's
downloads. The export writes the branch's real name; the import still accepts
the original file, because a mapping profile is what reconciles the two.
"""

from __future__ import annotations

import csv
import io

from .models import BranchStock, Product, ProductBarcode

PRODUCT_TEMPLATE_HEADERS = [
    'Código de barras',
    'Código',
    'Nombre',
    'Descripción',
    'Precio de venta',
    'Categoría',
]

PRODUCT_TEMPLATE_HELP = [
    'Se admite un EAN, un UPC o el código interno de la tienda.',
    'Código interno imprimible. Se registra como código de barras interno.',
    'Obligatorio.',
    'Opcional. Si se deja vacío no se borra la descripción existente.',
    'Obligatorio para productos nuevos. Usa punto decimal: 149.90',
    'Debe existir, salvo que se active «crear las que falten».',
]


def _workbook():
    import openpyxl
    return openpyxl.Workbook()


def _autosize(sheet, widths):
    from openpyxl.utils import get_column_letter
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def product_template_bytes() -> bytes:
    """
    A blank, VALID product template.

    Deliberately not a copy of the owner's 18-column sheet. That file carries
    two rows of headers, twenty data validations openpyxl cannot read back, and
    ten columns this catalogue has nowhere to put — including tax fields for
    electronic invoicing this platform does not do. Reproducing it would hand
    people a template whose columns are mostly ignored.

    The importer still reads that file. This is what we ASK for; that is what we
    ACCEPT.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = _workbook()
    sheet = workbook.active
    sheet.title = 'Productos'

    header_fill = PatternFill('solid', fgColor='1F2937')
    for column, title in enumerate(PRODUCT_TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center')
    for column, note in enumerate(PRODUCT_TEMPLATE_HELP, start=1):
        cell = sheet.cell(row=2, column=column, value=note)
        cell.font = Font(size=8, italic=True, color='6B7280')
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    sheet.cell(row=3, column=1, value='7751234567890')
    sheet.cell(row=3, column=2, value='C000001')
    sheet.cell(row=3, column=3, value='Cable Lightning a USB')
    sheet.cell(row=3, column=4, value='Cable de 1 metro')
    sheet.cell(row=3, column=5, value=49.90)
    sheet.cell(row=3, column=6, value='Accesorios')

    sheet.freeze_panes = 'A3'
    _autosize(sheet, [20, 16, 42, 40, 16, 20])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def inventory_export_bytes(*, company, branches, include_quantities: bool) -> bytes:
    """
    The catalogue with one column per branch, for counting and re-uploading.

    The header row is `CODIGO · CODIGO EAN · NOMBRE`, matching the shape the
    importer already recognises, plus one `<branch name>` column per branch. The
    result re-imports through the same detector as any other file.
    """
    from openpyxl.styles import Font, PatternFill

    branches = list(branches)
    workbook = _workbook()
    sheet = workbook.active
    sheet.title = 'Inventario'

    headers = ['CODIGO', 'CODIGO EAN', 'NOMBRE'] + [f'ALMACEN {b.name}' for b in branches]
    header_fill = PatternFill('solid', fgColor='1F2937')
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    products = list(
        Product.objects.filter(company=company, is_active=True).order_by('name')
    )
    codes: dict[int, dict[str, str]] = {}
    for product_id, code, symbology in ProductBarcode.objects.filter(
        company=company, is_active=True,
    ).values_list('product_id', 'code', 'symbology'):
        slot = codes.setdefault(product_id, {})
        if symbology == ProductBarcode.INTERNAL:
            slot.setdefault('internal', code)
        else:
            slot.setdefault('external', code)

    quantities = {
        (b, p): q
        for b, p, q in BranchStock.objects
        .filter(branch__in=branches)
        .values_list('branch_id', 'product_id', 'quantity')
    }

    for index, product in enumerate(products, start=2):
        slot = codes.get(product.pk, {})
        # Written as TEXT, always. A code stored as a number is how the owner's
        # own file lost the ability to keep a leading zero, and re-exporting the
        # same trap would repeat the damage on every round trip.
        sheet.cell(row=index, column=1, value=slot.get('internal', '')).number_format = '@'
        sheet.cell(row=index, column=2, value=slot.get('external', '')).number_format = '@'
        sheet.cell(row=index, column=3, value=product.name)
        for offset, branch in enumerate(branches):
            cell = sheet.cell(row=index, column=4 + offset)
            if include_quantities:
                cell.value = quantities.get((branch.pk, product.pk), 0)
            # else: left EMPTY, which the importer reads as "do not change" —
            # so a sheet that comes back with only twelve lines filled in moves
            # exactly twelve stocks.

    sheet.freeze_panes = 'A2'
    _autosize(sheet, [16, 18, 46] + [18] * len(branches))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def error_report_csv(job) -> str:
    """
    The rows that failed, as CSV (§63).

    CSV rather than XLSX because this file exists to be read next to the
    original: it is opened, scanned, and thrown away. UTF-8 with a BOM, because
    without it Excel on Windows renders every accented product name as mojibake
    and the report becomes unreadable in the one program that will open it.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Hoja', 'Fila', 'Identificador', 'Producto', 'Error', 'Advertencia'])
    for row in job.rows.exclude(errors=[]).order_by('row_number', 'pk'):
        data = row.normalized_data or {}
        writer.writerow([
            row.sheet_name, row.row_number, row.match_key,
            data.get('name', ''),
            ' · '.join(row.errors or []),
            ' · '.join(row.warnings or []),
        ])
    return '﻿' + buffer.getvalue()
